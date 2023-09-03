import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
import tempfile
import os
import locale
from io import BytesIO
from xlsx2html import xlsx2html
import pdfkit
import yagmail
# Streamlit app
import numpy as np
import os
from openpyxl.styles import Font, PatternFill, Border, Side
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def main():
    sender_email = '' #put your email address here
    sender_password = '' #put your password here (app_password if you have 2fa turned on)
    st.title('Invoice Generator')
    client_info = pd.read_excel('Info_clients.xlsx')

    # Upload Excel file
    uploaded_file = st.file_uploader('Upload hier je urenoverzicht', type=['xlsx'])


    # Function to process the uploaded Excel file
    def process_excel_file(uploaded_file,client_info):
        df = pd.read_excel(uploaded_file, header=16)
        #st.dataframe(df)
        df['Naam Locatie'].replace('', np.nan, inplace=True)
        df.dropna(subset=['Naam Locatie'], inplace=True)
        df['Medewerker'][0] = df['Medewerker'][0].lower()
        client_info.iloc[:,1] = [x.lower() for x in client_info.iloc[:,1]]
        Uurtarief = st.session_state.Uurtarief = client_info.iloc[:,10][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        
        # User input for KM_vergoeding
        Km_vergoeding = st.session_state.km_vergoeding = client_info.iloc[:,11][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        mail = client_info.iloc[:,12][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        # Variables
        Rekeningummer = client_info.iloc[:,8][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        KVK = client_info.iloc[:,7][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        Btw_nummer = client_info.iloc[:,9][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]
        Bedrijfsnaam = client_info.iloc[:,4][list(client_info.iloc[:,1]).index(df['Medewerker'][0])]

     

          # Formatting 'datum' column
        df['datum'] = pd.to_datetime(df['datum']).dt.strftime("%d-%m-%Y")
        # Formatting 'Begintijd' column
        df['Begintijd'] = pd.to_datetime(df['Begintijd'], format='%H:%M:%S').dt.strftime('%H:%M')
        # Formatting 'Eindtijd' column
        df['Eindtijd'] = pd.to_datetime(df['Eindtijd'], format='%H:%M:%S').dt.strftime('%H:%M')




       # Create a dictionary to store user-provided kilometers per row
        km_input_dict = []
        default_km = 150
        distinct_locaties = df[df['Naam Locatie'] != 'Totaal']['Naam Locatie']

        # Create an input widget for each unique locatie
        for locatie in range(len(distinct_locaties)):
            km_input = st.number_input(f"{distinct_locaties[locatie]} - Noteer kilometers: {df['datum'][locatie]}", min_value=0, value=default_km, max_value=200, key=locatie)
            km_input_dict.append(km_input)
       
        for x in range(len(df)-1):
            df.loc[x,'KM'] = km_input_dict[x]
       
        df.dropna()
        df.drop(['Medewerker', 'Dag','Slaap of Waak?', 'Totaal'], axis=1)

        # Define the desired column order
        desired_order = ['datum', 'Naam Locatie', 1, 1.22, 1.38, 1.44, 1.49, 1.6, 'Begintijd', 'Eindtijd', 'uren maal toeslag','KM']

        # Reindex the dataframe with the desired column order
        df = df.reindex(columns=desired_order)

        #creating total hours columns
        df['uren maal toeslag'] = df[1]*1 + df[1.22]*1.22 + df[1.38]*1.38 + df[1.44]*1.44 + df[1.49]*1.49 + df[1.6]*1.6

        # Rounding 'uren totaal incl %' column to 2 decimal places
        df['uren maal toeslag'] = round(df['uren maal toeslag'], 2)

    
        for i in range(len(df) - 1):
            current_row = df.iloc[i]
            previous_row = df.iloc[i - 1]
            
            if current_row['datum'] == previous_row['datum']:
                    df.at[i, 'KM'] = 0


        df = df.drop(df.index[-1])
    
    
        # Calculate the sum of 'uren maal toeslag' and 'KM' columns
        total_row = pd.DataFrame({
            'uren maal toeslag': df['uren maal toeslag'].sum(),
            'KM': df['KM'].sum(),
            'Eindtijd': 'Totaal'
        }, index=['Total'])

        # Append the total row to the input_file dataframe
        df = df.append(total_row)

        # Reset the index of the dataframe
        df.reset_index(drop=True, inplace=True)
        df = df.fillna('')

        # Set the locale to Dutch (Netherlands)
        locale.setlocale(locale.LC_TIME, 'nl_NL.UTF-8')

        # Get the current date
        current_date = datetime.now()

        # Extract year and month from the current date
        year = current_date.year
        month_number = current_date.month

        # Get the full month name based on the month number
        month_name = current_date.strftime('%B')

        # Generate the fac_number combining year and current month number
        fac_number = f"{year}{month_number:02d}"

        # Get the whole date today in your desired format
        date_today = current_date.strftime('%Y-%m-%d')
    
        ## Calculate the values for the table
        Totale_uren = total_row['uren maal toeslag'].iloc[0]
        Uren_maal_tarief = Totale_uren * Uurtarief
        Totale_km = total_row['KM'].iloc[0]
        KM_maal_tarief = Totale_km * Km_vergoeding
        sub_totaal = Uren_maal_tarief + KM_maal_tarief
        
        return mail, df, Totale_uren, Uurtarief, Uren_maal_tarief, Totale_km, Km_vergoeding, KM_maal_tarief, sub_totaal, Bedrijfsnaam, KVK, Rekeningummer, Btw_nummer, month_name, fac_number, date_today


    if uploaded_file is None:
        st.write("")

    else:
        name = pd.read_excel(uploaded_file, header=16)
        name = name['Medewerker'][0]
        
        # Process the uploaded file and get the variables
        mail, df, Totale_uren, Uurtarief, Uren_maal_tarief, Totale_km, Km_vergoeding, KM_maal_tarief, sub_totaal, Bedrijfsnaam, KVK, Rekeningummer, Btw_nummer, month_name, fac_number, date_today = process_excel_file(uploaded_file,client_info)
        df0 = pd.read_excel(uploaded_file, header=16)
        df0['Medewerker'][0] = df0['Medewerker'][0].lower()
        
        df.rename(columns={
            'datum': 'Datum',
            'Naam Locatie': 'Locatie',
            1: '1',
            1.22: '1.22',
            1.38: '1.38',
            1.44: '1.44',
            1.49: '1.49',
            1.6: '1.6',
            'uren maal toeslag': 'Uren incl. toeslag',
            'KM': 'Afstand (km)'
        }, inplace=True)
        
        # Define the text information
        text_info = [
            ["Factuur"],
            [""],
            ["Allround Care", "", "", "", "", "", "", "", "", Bedrijfsnaam],
            ["T.a.v. Administratie", "", "", "", "", "", "", "", "", client_info.iloc[:, 5][list(client_info.iloc[:, 1]).index(df0['Medewerker'][0])] ],
            ["Wiersedreef 22", "", "", "", "", "", "", "", "", client_info.iloc[:, 6][list(client_info.iloc[:, 1]).index(df0['Medewerker'][0])] ],
            ["3433 ZX Nieuwegein"],
            [""],
            [""],
            ["Maand: " + month_name, "", "", "", "", "", "", "", "", "KVK: "+str(KVK)],
            ["Factuurnummer: ", fac_number,"", "", "", "", "", "", "", "Rekeningummer: " +str(Rekeningummer)],
            ["Factuurdatum: ", date_today,"", "", "", "", "", "", "", "Btw nummer: " + str(Btw_nummer)]
        ]
        
        # Define the data for the row
        data = [
            [''],
            [''],
            ['','','','','','','','Uurtarief maal uren','', Totale_uren, Uurtarief, '€' + str(round(Uren_maal_tarief,2)) ],
            ['','','','','','','','Totaal kilometervergoeding','', Totale_km, Km_vergoeding, KM_maal_tarief],
            ['','','','','','','','Sub-totaal','', '', '', '€' + str(round(sub_totaal,2))],
            ['','','','','','','','BTW%','', '21%', '', '€' + str(round(0.21 * sub_totaal,2))],
            ['','','','','','','','FACTUUR BEDRAG','', '', '', '€' + str(round(1.21 * sub_totaal,2))],
            ['','','','',],
            ['','','','','','','','Rekeningnummer','', Rekeningummer, '', ''],
            [""],
            ['Wij verzoeken u vriendelijk het bovenstaande bedrag binnen 30 dagen over te maken op bovenstaand rekeningnummer onder vermelding van het factuurnummer.', '', '', ''],
        ]
         # Create the dataframe
        df2 = pd.DataFrame(data)

        
        
        for i in range(len(df)):
            #st.write(df.iloc[:,2][i])
            if isinstance(df.iloc[:,2][i],str) == False:
             df.iloc[:,2][i] = round(float(df.iloc[:,2][i]),3)
            if isinstance(df.iloc[:,3][i],str) == False:    
             df.iloc[:,3][i] = round(float(df.iloc[:,3][i]),3)
            if isinstance(df.iloc[:,4][i],str) == False:
             df.iloc[:,4][i] = round(float(df.iloc[:,4][i]),3)
            if isinstance(df.iloc[:,5][i],str) == False:
             df.iloc[:,5][i] = round(float(df.iloc[:,5][i]),3)
            if isinstance(df.iloc[:,6][i],str) == False:
             df.iloc[:,6][i] = round(float(df.iloc[:,6][i]),3)
            if isinstance(df.iloc[:,7][i],str) == False:
             df.iloc[:,7][i] = round(float(df.iloc[:,7][i]),3)
            if isinstance(df.iloc[:,10][i],str) == False:    
             df.iloc[:,10][i] = round(float(df.iloc[:,10][i]),3)
        # Show the DataFrame with the calculated total amounts
        st.write('preview factuur')
        st.dataframe(df)

        st.write('Factuurbedrag totaal:', f'€{1.21 * sub_totaal:.2f}')
    
        #if st.button('Download de factuur'):
        # Create an ExcelWriter object
        excel_file_path = 'Factuur.xlsx'
        
        # Create an ExcelWriter with the openpyxl engine
        writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')

         # Write the dataframes to the Excel file
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=True, startrow=len(text_info) + 1)
        df2.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=len(text_info) + 2 + len(df))

        # Get the workbook
        workbook = writer.book

        # Write the text information to the sheet
        worksheet = writer.sheets['Sheet1']
        for row_idx, row_data in enumerate(text_info, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)     

        # Remove gridlines for the entire sheet
        worksheet.sheet_view.showGridLines = False

        
        worksheet.column_dimensions['A'].width = 11
        worksheet.column_dimensions['B'].width = 13
        worksheet.column_dimensions['C'].width = 5
        worksheet.column_dimensions['D'].width = 5
        worksheet.column_dimensions['E'].width = 5
        worksheet.column_dimensions['F'].width = 5
        worksheet.column_dimensions['G'].width = 5
        worksheet.column_dimensions['H'].width = 5
        worksheet.column_dimensions['I'].width = 8
        worksheet.column_dimensions['J'].width = 8
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 12

        # Set the font and background color of the headers from row 14
        for row in worksheet.iter_rows(min_row=13, max_row=13):
            for cell in row:
                cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type="solid")

        # Set the font and background color for cells in columns J, K, and L for row 32 (totaal uren etc.)
        for col_letter in ['J', 'K', 'L']:
            cell = worksheet[f'{col_letter}32']
            cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type="solid")
            cell.font = Font(bold=True)  # Change font color to white and make it bold

        # Set the font style to bold for cells in rows 39 (H39 to L39)
        for col_letter in ['H', 'I', 'J', 'K', 'L']:
            cell = worksheet[f'{col_letter}39']
            cell.font = Font(bold=True)    

        # Set the font size to 8 for cells in column A from row 1 to 42
        # Set the font size to 8 for cell A43
        cell = worksheet['A43']
        cell.font = Font(size=8)  


            

        # Write the text_info to the sheet and format the cell with "Factuur"
        for row_idx, row_data in enumerate(text_info, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                if cell_value == "Factuur":
                    cell.font = Font(color='1F497D', bold=True, size=28)

       
        # # Set the font size for row 39, column 8 (assuming you want to set the font for cell H39)
        # row_idx = 39
        # col_idx = 1
        # # Get the cell
        # cell = worksheet.cell(row=row_idx, column=col_idx)
        # # Create a Font object with size 8
        # font = Font(size=8)
        # # Set the font for the cell
        # cell.font = font
        
        row_idx = 34
        col_idx = 12
        # Get the cell
        cell = worksheet.cell(row=row_idx, column=col_idx)
        # Create a Font object with size 8
        font = Font(bold=True)
        # Set the font for the cell
        cell.font = font
                    

            
        # Add a border to row 13 from A1 to M39
        # Define the border style
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add a bottom border to cells in row 43 (A43:L43)
    for col_idx in range(1, 13):  # Columns A to L
        cell = worksheet.cell(row=43, column=col_idx)
        cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))


        # Add a right border to cells in column L (L1:L43)
    for row_idx in range(1, 44):  # Rows 1 to 43
        cell = worksheet.cell(row=row_idx, column=12)  # Column L
        cell.border = Border(right=Side(style='thin'))
        
   #Add a right border to cells in column L (L1:L43)
    for row_idx in range(1, 44):  # Rows 1 to 43
        cell = worksheet.cell(row=row_idx, column=12)  # Column L
        cell.border = Border(right=Side(style='thin'))

    # Add borders to all cells in row 13 (A13:L13)
    for col_idx in range(1, 14):  # Columns A to L
        cell = worksheet.cell(row=13, column=col_idx)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))                                    
         
        # List of cell ranges to merge
        cell_ranges_to_merge = ['H20:I20', 'H30:I30', 'H31:I31', 'H33:I33', 'H34:I34']

        # Merge the specified cell ranges
        for cell_range in cell_ranges_to_merge:
            worksheet.merge_cells(cell_range)
                
              
       # Save the workbook
        writer.save()

        # Load the workbook after saving
        workbook = openpyxl.load_workbook('Factuur.xlsx')

        # Select the sheet to modify
        sheet = workbook['Sheet1']


        
        with open(excel_file_path, "rb") as file:
          btn=st.download_button(
          label="Download hier je factuur (xlsx)",
          data=file,
          file_name=excel_file_path,
          mime="application/octet-stream"
          )


        xlsx2html('Factuur.xlsx', 'file.html')
        

        config = pdfkit.configuration(wkhtmltopdf="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe")
        pdfkit.from_file("file.html", "Factuur.pdf",configuration=config)
        

        with open("Factuur.pdf", "rb") as file:
           btn=st.download_button(
           label="Download hier je factuur (pdf)",
           data=file,
           file_name="Factuur.pdf",
           mime="application/octet-stream"
           )
        if st.button('Verzend mail'):
         receiver = mail
         body = """Geachte, 

                In de bijlage treft u de factuur. 

                Voor vragen en/of opmerkingen kunt u contact met mij opnemen.

                Met vriendelijke groet,

                Invoice Generator"""
         filename = ["Factuur.pdf","Factuur.xlsx"]
         try:   
           yag = yagmail.SMTP(sender_email,sender_password)
           yag.send(
                 to=receiver,
                 subject='Factuur '+ month_name + ' ' + name ,
                 contents=body, 
                 attachments=filename,
                )
           st.success(f'E-mail succesvol verzonden!')
         except Exception as e:
             st.write(e)
if __name__ == '__main__':
    main()

